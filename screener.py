#!/usr/bin/env python3
"""
VWRP (Vanguard FTSE All-World UCITS ETF) stock screener.

Pulls fundamentals for the largest constituents of the FTSE All-World universe
and scores each on valuation, balance-sheet health, revenue growth and risk.

Outputs:
  output/vwrp_screen_<date>.xlsx   sortable/filterable spreadsheet
  output/dashboard.html            interactive dashboard

Usage:
  python screener.py                  # top 150 by market cap
  python screener.py --top 100        # top 100
  python screener.py --refresh        # ignore cache, re-fetch everything
  python screener.py --max-pe 25 --max-de 100   # only show stocks passing filters
"""

import argparse
import csv
import json
import math
import os
import re
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta

import pandas as pd
import yfinance as yf

BASE = os.path.dirname(os.path.abspath(__file__))
UNIVERSE = os.path.join(BASE, "universe.csv")
HOLDINGS = os.path.join(BASE, "holdings.csv")
CACHE = os.path.join(BASE, ".cache")
OUTDIR = os.path.join(BASE, "output")
CACHE_TTL_HOURS = 24


# --------------------------------------------------------------------------
# Universe
# --------------------------------------------------------------------------

def load_holdings(pool):
    """VWRP's real constituents, heaviest first, straight from Vanguard.

    We keep the whole fund on file but only fetch the heaviest `pool` of them:
    a holding at rank 2000 carries 0.003% and would need to grow 35x to reach
    the top 150, so fetching it nightly buys nothing."""
    rows = []
    with open(HOLDINGS) as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or line.startswith("ticker,"):
                continue
            parts = line.split(",")
            if len(parts) >= 3 and parts[0]:
                rows.append({"ticker": parts[0], "region": parts[1],
                             "vanguard_weight": float(parts[2]),
                             "holding_name": parts[3] if len(parts) > 3 else None})
    rows.sort(key=lambda r: -r["vanguard_weight"])
    return rows[:pool] if pool else rows


def load_universe():
    rows = []
    with open(UNIVERSE) as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or line.startswith("ticker,"):
                continue
            parts = [p.strip() for p in line.split(",")]
            if len(parts) >= 2 and parts[0]:
                rows.append({"ticker": parts[0], "region": parts[1]})
    return rows


# --------------------------------------------------------------------------
# Fetching
# --------------------------------------------------------------------------

def cache_path(ticker):
    safe = ticker.replace("/", "_").replace(".", "_")
    return os.path.join(CACHE, safe + ".json")


def read_cache(ticker):
    p = cache_path(ticker)
    if not os.path.exists(p):
        return None
    age = datetime.now() - datetime.fromtimestamp(os.path.getmtime(p))
    if age > timedelta(hours=CACHE_TTL_HOURS):
        return None
    try:
        with open(p) as fh:
            return json.load(fh)
    except Exception:
        return None


def write_cache(ticker, data):
    os.makedirs(CACHE, exist_ok=True)
    try:
        with open(cache_path(ticker), "w") as fh:
            json.dump(data, fh)
    except Exception:
        pass


WANTED_INFO = [
    "shortName", "longName", "sector", "industry", "country", "currency",
    "financialCurrency", "marketCap", "trailingPE", "forwardPE", "priceToBook",
    "priceToSalesTrailing12Months", "debtToEquity", "totalDebt", "totalCash",
    "revenueGrowth", "earningsGrowth", "returnOnEquity", "profitMargins",
    "operatingMargins", "grossMargins", "beta", "freeCashflow",
    "operatingCashflow", "ebitda", "totalRevenue", "currentRatio",
    "quickRatio", "dividendYield", "payoutRatio", "trailingEps",
    "enterpriseValue", "enterpriseToEbitda", "recommendationKey",
    "numberOfAnalystOpinions", "targetMeanPrice", "currentPrice",
    "fiftyTwoWeekHigh", "fiftyTwoWeekLow",
]


def fetch_one(ticker, use_cache=True, attempts=3):
    """Fetch info + multi-year revenue for one ticker.

    Yahoo silently drops fields (marketCap especially) when it throttles a
    burst, so retry with backoff and fall back to shares x price."""
    if use_cache:
        cached = read_cache(ticker)
        if cached is not None:
            return cached

    rec = {"ticker": ticker, "error": None}
    info = {}
    for attempt in range(attempts):
        try:
            info = yf.Ticker(ticker).info or {}
        except Exception as exc:
            rec["error"] = str(exc)[:200]
            info = {}
        if info.get("marketCap"):
            rec["error"] = None
            break
        if "Quote not found" in (rec.get("error") or ""):
            break  # genuinely bad symbol, no point retrying
        if attempt < attempts - 1:
            time.sleep(1.5 * (attempt + 1))

    try:
        t = yf.Ticker(ticker)
        for k in WANTED_INFO:
            rec[k] = info.get(k)

        if not rec.get("marketCap"):
            shares = info.get("sharesOutstanding") or info.get("impliedSharesOutstanding")
            price = (info.get("currentPrice") or info.get("regularMarketPrice")
                     or info.get("previousClose"))
            if shares and price:
                rec["marketCap"] = shares * price
                rec["marketCap_estimated"] = True

        # Multi-year revenue history for trend analysis
        revs = []
        try:
            stmt = t.income_stmt
            if stmt is not None and not stmt.empty:
                row = None
                for label in ("Total Revenue", "TotalRevenue", "Operating Revenue"):
                    if label in stmt.index:
                        row = stmt.loc[label]
                        break
                if row is not None:
                    for col in sorted(stmt.columns, reverse=True):
                        val = row.get(col)
                        if val is not None and not (isinstance(val, float) and math.isnan(val)):
                            revs.append({"year": int(pd.Timestamp(col).year),
                                         "revenue": float(val)})
        except Exception:
            pass
        rec["revenue_history"] = revs

        if not rec.get("marketCap"):
            rec["error"] = "no market cap"
    except Exception as exc:
        rec["error"] = str(exc)[:200]

    # Only cache successes — a rate-limited miss must not poison the next run.
    if rec.get("marketCap"):
        write_cache(ticker, rec)
    return rec


def fetch_all(universe, use_cache=True, workers=8, delay=0.0):
    results = []
    total = len(universe)
    done = 0

    def job(ticker):
        if delay:
            time.sleep(delay)
        return fetch_one(ticker, use_cache)

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(job, u["ticker"]): u for u in universe}
        for fut in as_completed(futures):
            u = futures[fut]
            done += 1
            try:
                rec = fut.result()
            except Exception as exc:
                rec = {"ticker": u["ticker"], "error": str(exc)[:200]}
            rec["region"] = u["region"]
            for k in ("vanguard_weight", "holding_name"):
                if u.get(k) is not None:
                    rec[k] = u[k]
            results.append(rec)
            if done % 20 == 0 or done == total:
                print(f"  fetched {done}/{total}", file=sys.stderr)
    return results


def fetch_fx(currencies):
    """Spot rates to USD. Yahoo reports marketCap in each listing's own
    currency, so ranking without this puts KRW/JPY names at the top."""
    rates = {"USD": 1.0}
    # Pence/cents-quoted listings still report marketCap in the major unit.
    minor = {"GBp": "GBP", "ZAc": "ZAR", "ILA": "ILS"}
    needed = set()
    for c in currencies:
        if not c:
            continue
        needed.add(minor.get(c, c))
    needed.discard("USD")

    for cur in sorted(needed):
        rate = None
        for pair in (f"{cur}USD=X", f"USD{cur}=X"):
            try:
                hist = yf.Ticker(pair).history(period="5d")
                if hist is not None and not hist.empty:
                    px = float(hist["Close"].iloc[-1])
                    if px > 0:
                        rate = px if pair.startswith(cur) else 1.0 / px
                        break
            except Exception:
                continue
        if rate:
            rates[cur] = rate
        else:
            print(f"  warning: no FX rate for {cur}, leaving unconverted",
                  file=sys.stderr)
            rates[cur] = 1.0
    for src, dst in minor.items():
        if dst in rates:
            rates[src] = rates[dst]
    return rates


def fetch_price_risk(tickers):
    """Bulk-download 1y prices to compute realised volatility and drawdown."""
    out = {}
    print("Downloading price history for volatility/drawdown...", file=sys.stderr)
    try:
        data = yf.download(tickers, period="1y", interval="1d",
                           auto_adjust=True, progress=False, threads=True)
    except Exception as exc:
        print(f"  price download failed: {exc}", file=sys.stderr)
        return out

    if data is None or data.empty:
        return out

    close = data["Close"] if "Close" in data.columns.get_level_values(0) else data
    if isinstance(close, pd.Series):
        close = close.to_frame(tickers[0])

    for tk in close.columns:
        s = close[tk].dropna()
        if len(s) < 60:
            continue
        ret = s.pct_change().dropna()
        vol = float(ret.std() * math.sqrt(252) * 100)
        peak = s.cummax()
        dd = float(((s / peak) - 1).min() * 100)
        perf = float((s.iloc[-1] / s.iloc[0] - 1) * 100)
        last = float(s.iloc[-1])
        prev = float(s.iloc[-2]) if len(s) > 1 else None
        out[tk] = {"volatility_1y_pct": vol, "max_drawdown_1y_pct": dd,
                   "return_1y_pct": perf, "price": last,
                   "day_change_pct": ((last / prev - 1) * 100) if prev else None}
    return out


# --------------------------------------------------------------------------
# Derived metrics
# --------------------------------------------------------------------------

def revenue_trend(history):
    """Return YoY growth list, 3y CAGR, and a trend label."""
    out = {"rev_yoy_1": None, "rev_yoy_2": None, "rev_yoy_3": None,
           "rev_cagr_3y": None, "rev_trend": None, "rev_years": 0}
    if not history:
        return out
    hist = sorted(history, key=lambda r: r["year"], reverse=True)
    out["rev_years"] = len(hist)

    yoys = []
    for i in range(min(3, len(hist) - 1)):
        newer, older = hist[i]["revenue"], hist[i + 1]["revenue"]
        if older and older > 0:
            yoys.append((newer / older - 1) * 100)
        else:
            yoys.append(None)
    for i, v in enumerate(yoys):
        out[f"rev_yoy_{i+1}"] = v

    if len(hist) >= 2:
        newest, oldest = hist[0]["revenue"], hist[-1]["revenue"]
        years = hist[0]["year"] - hist[-1]["year"]
        if oldest and oldest > 0 and years > 0:
            out["rev_cagr_3y"] = ((newest / oldest) ** (1 / years) - 1) * 100

    valid = [v for v in yoys if v is not None]
    if len(valid) >= 2:
        recent, prior = valid[0], valid[1]
        if recent > 0 and prior > 0:
            out["rev_trend"] = "Accelerating" if recent > prior + 2 else (
                "Decelerating" if recent < prior - 2 else "Steady")
        elif recent > 0 >= prior:
            out["rev_trend"] = "Recovering"
        elif recent <= 0 < prior:
            out["rev_trend"] = "Contracting"
        else:
            out["rev_trend"] = "Contracting"
    elif len(valid) == 1:
        out["rev_trend"] = "Growing" if valid[0] > 0 else "Contracting"
    return out


def scale(value, low, high):
    """Map value into 0-100 risk points, clamped."""
    if value is None:
        return None
    if high == low:
        return 0.0
    pts = (value - low) / (high - low) * 100
    return max(0.0, min(100.0, pts))


def risk_score(r):
    """Composite 0-100 risk score. Higher = riskier. Weights renormalised
    over whatever components have data, so partial coverage still scores."""
    comps = {}

    # 1. Leverage
    de = r.get("debtToEquity")
    lev = scale(de, 0, 250) if de is not None else None
    nd_ebitda = r.get("net_debt_to_ebitda")
    if nd_ebitda is not None:
        nd_pts = scale(nd_ebitda, 0, 5)
        lev = nd_pts if lev is None else (lev * 0.6 + nd_pts * 0.4)
    if lev is not None:
        comps["leverage"] = (lev, 0.25)

    # 2. Volatility
    vol_parts = []
    beta = r.get("beta")
    if beta is not None:
        vol_parts.append(scale(abs(beta), 0.4, 2.0))
    v1 = r.get("volatility_1y_pct")
    if v1 is not None:
        vol_parts.append(scale(v1, 12, 60))
    dd = r.get("max_drawdown_1y_pct")
    if dd is not None:
        vol_parts.append(scale(abs(dd), 8, 55))
    if vol_parts:
        comps["volatility"] = (sum(vol_parts) / len(vol_parts), 0.25)

    # 3. Valuation
    pe = r.get("forwardPE") or r.get("trailingPE")
    val_parts = []
    if pe is not None:
        val_parts.append(100.0 if pe < 0 else scale(pe, 10, 55))
    pb = r.get("priceToBook")
    if pb is not None and pb > 0:
        val_parts.append(scale(pb, 1, 12))
    if val_parts:
        comps["valuation"] = (sum(val_parts) / len(val_parts), 0.20)

    # 4. Growth quality
    g_parts = []
    cagr = r.get("rev_cagr_3y")
    if cagr is not None:
        g_parts.append(100 - scale(cagr, -10, 30))
    yoy = r.get("rev_yoy_1")
    if yoy is not None:
        g_parts.append(100 - scale(yoy, -15, 35))
    if r.get("rev_trend") in ("Contracting", "Decelerating"):
        g_parts.append(75.0)
    elif r.get("rev_trend") in ("Accelerating",):
        g_parts.append(20.0)
    if g_parts:
        comps["growth"] = (sum(g_parts) / len(g_parts), 0.20)

    # 5. Profitability / cash generation
    p_parts = []
    roe = r.get("returnOnEquity")
    if roe is not None:
        p_parts.append(100 - scale(roe * 100, -5, 35))
    pm = r.get("profitMargins")
    if pm is not None:
        p_parts.append(100 - scale(pm * 100, -5, 30))
    fcf = r.get("freeCashflow")
    if fcf is not None:
        p_parts.append(85.0 if fcf < 0 else 20.0)
    if p_parts:
        comps["profitability"] = (sum(p_parts) / len(p_parts), 0.10)

    if not comps:
        return None, None, 0
    tw = sum(w for _, w in comps.values())
    score = sum(v * w for v, w in comps.values()) / tw
    coverage = tw / 1.00 * 100
    band = ("Low" if score < 35 else
            "Moderate" if score < 50 else
            "Elevated" if score < 65 else "High")
    return round(score, 1), band, round(coverage)


def _median(vals):
    if not vals:
        return None
    v = sorted(vals)
    m = len(v) // 2
    return v[m] if len(v) % 2 else (v[m - 1] + v[m]) / 2


def sector_benchmarks(rows, min_n=5):
    """Median multiple per sector, so 'cheap' means cheap for that industry.

    Without this, banks and insurers sweep the top of any value ranking --
    they trade on structurally low P/E and P/B by convention, which says
    nothing about whether one bank is better value than another."""
    buckets = {}
    for r in rows:
        sec = r.get("sector") or "Unknown"
        b = buckets.setdefault(sec, {"pe": [], "pb": [], "ev": []})
        pe = r.get("forwardPE") or r.get("trailingPE")
        if pe and pe > 0:
            b["pe"].append(pe)
        if r.get("priceToBook") and r["priceToBook"] > 0:
            b["pb"].append(r["priceToBook"])
        if r.get("enterpriseToEbitda") and r["enterpriseToEbitda"] > 0:
            b["ev"].append(r["enterpriseToEbitda"])
    return {sec: {k: (_median(v) if len(v) >= min_n else None)
                  for k, v in b.items()} for sec, b in buckets.items()}


def opportunity_score(r, bench=None):
    """Composite 0-100, HIGHER = screens better. Blends cheapness, growth and
    balance-sheet quality, then applies a drag for price risk.

    This is a mechanical ranking of reported numbers, not a view on any
    company. It deliberately rewards low multiples, so it will surface value
    traps and cyclicals at peak earnings alongside genuine bargains -- read the
    Flags column and the risk rating before trusting the order."""
    comps = {}

    # 1. Value -- cheap multiples score high, judged against the stock's own
    # sector where we have enough peers to form a median.
    sec_bench = (bench or {}).get(r.get("sector") or "Unknown", {})

    def value_pts(val, key, abs_lo, abs_hi):
        if val is None or val <= 0:
            return None
        med = sec_bench.get(key)
        if med:
            return 100 - scale(val / med, 0.5, 1.8)   # 0.5x sector = cheap
        return 100 - scale(val, abs_lo, abs_hi)

    v = []
    pe = r.get("forwardPE") or r.get("trailingPE")
    if pe is not None and pe <= 0:
        v.append(0.0)                                  # loss-making
    else:
        v.append(value_pts(pe, "pe", 8, 45))
    v.append(value_pts(r.get("priceToBook"), "pb", 0.8, 10))
    v.append(value_pts(r.get("enterpriseToEbitda"), "ev", 5, 30))
    v = [x for x in v if x is not None]
    if v:
        comps["value"] = (sum(v) / len(v), 0.30)

    # 2. Growth -- expanding revenue scores high
    g = []
    if r.get("rev_yoy_1") is not None:
        g.append(scale(r["rev_yoy_1"], -10, 30))
    if r.get("rev_cagr_3y") is not None:
        g.append(scale(r["rev_cagr_3y"], -5, 25))
    trend_pts = {"Accelerating": 90, "Recovering": 70, "Steady": 60,
                 "Growing": 60, "Decelerating": 30, "Contracting": 10}
    if r.get("rev_trend") in trend_pts:
        g.append(float(trend_pts[r["rev_trend"]]))
    if g:
        comps["growth"] = (sum(g) / len(g), 0.30)

    # 3. Quality -- profitable, cash-generative, not over-levered
    q = []
    if r.get("returnOnEquity") is not None:
        q.append(scale(r["returnOnEquity"] * 100, 0, 30))
    if r.get("profitMargins") is not None:
        q.append(scale(r["profitMargins"] * 100, 0, 25))
    if r.get("debtToEquity") is not None:
        q.append(100 - scale(r["debtToEquity"], 0, 200))
    if r.get("freeCashflow") is not None:
        q.append(20.0 if r["freeCashflow"] < 0 else 80.0)
    if r.get("currentRatio") is not None:
        q.append(scale(r["currentRatio"], 0.5, 2.5))
    if q:
        comps["quality"] = (sum(q) / len(q), 0.25)

    # 4. Stability -- low volatility and shallow drawdowns score high
    s = []
    if r.get("beta") is not None:
        s.append(100 - scale(abs(r["beta"]), 0.4, 2.0))
    if r.get("volatility_1y_pct") is not None:
        s.append(100 - scale(r["volatility_1y_pct"], 12, 60))
    if r.get("max_drawdown_1y_pct") is not None:
        s.append(100 - scale(abs(r["max_drawdown_1y_pct"]), 8, 55))
    if s:
        comps["stability"] = (sum(s) / len(s), 0.15)

    if not comps:
        return None, None
    tw = sum(w for _, w in comps.values())
    score = sum(val * w for val, w in comps.values()) / tw
    band = ("Strong" if score >= 62 else
            "Good" if score >= 52 else
            "Fair" if score >= 42 else "Weak")
    return round(score, 1), band


def health_flags(r):
    """Plain-English warnings on balance sheet / growth."""
    flags = []
    de = r.get("debtToEquity")
    if de is not None:
        if de > 200:
            flags.append("Very high debt/equity")
        elif de > 100:
            flags.append("High debt/equity")
    if r.get("net_debt_to_ebitda") is not None and r["net_debt_to_ebitda"] > 3:
        flags.append("Net debt >3x EBITDA")
    pe = r.get("trailingPE")
    if pe is not None and pe < 0:
        flags.append("Loss-making")
    elif pe is not None and pe > 50:
        flags.append("Rich valuation")
    if r.get("rev_trend") == "Contracting":
        flags.append("Revenue contracting")
    elif r.get("rev_trend") == "Decelerating":
        flags.append("Growth decelerating")
    if r.get("freeCashflow") is not None and r["freeCashflow"] < 0:
        flags.append("Negative free cash flow")
    if r.get("currentRatio") is not None and r["currentRatio"] < 1:
        flags.append("Current ratio <1")
    return flags


def build_rows(records, price_risk):
    rows = []
    for r in records:
        if r.get("error") and not r.get("marketCap"):
            continue
        mc = r.get("marketCap")
        if not mc:
            continue
        row_mc_usd = r.get("marketCap_usd") or mc

        row = dict(r)
        row["marketCap_usd"] = row_mc_usd
        row.update(price_risk.get(r["ticker"], {}))

        # Net debt / EBITDA
        td, tc, ebitda = r.get("totalDebt"), r.get("totalCash"), r.get("ebitda")
        if td is not None and ebitda and ebitda > 0:
            row["net_debt_to_ebitda"] = (td - (tc or 0)) / ebitda
        else:
            row["net_debt_to_ebitda"] = None

        row.update(revenue_trend(r.get("revenue_history")))

        score, band, cov = risk_score(row)
        row["risk_score"] = score
        row["risk_band"] = band
        row["data_coverage_pct"] = cov
        row["flags"] = "; ".join(health_flags(row))
        row["name"] = r.get("longName") or r.get("shortName") or r["ticker"]
        rows.append(row)

    bench = sector_benchmarks(rows)
    for row in rows:
        opp, opp_band = opportunity_score(row, bench)
        row["opportunity_score"] = opp
        row["opportunity_band"] = opp_band
    return rows


# --------------------------------------------------------------------------
# Output
# --------------------------------------------------------------------------

COLUMNS = [
    ("rank", "Rank"),
    ("ticker", "Ticker"),
    ("name", "Company"),
    ("price", "Price"),
    ("day_change_pct", "Day %"),
    ("currency", "Cur"),
    ("opportunity_score", "Opportunity score"),
    ("opportunity_band", "Opportunity"),
    ("risk_score", "Risk score"),
    ("risk_band", "Risk rating"),
    ("sector", "Sector"),
    ("region", "Region"),
    ("trailingPE", "P/E (trailing)"),
    ("forwardPE", "P/E (forward)"),
    ("priceToBook", "P/B"),
    ("debtToEquity", "Debt/Equity %"),
    ("net_debt_to_ebitda", "Net debt/EBITDA"),
    ("currentRatio", "Current ratio"),
    ("rev_yoy_1", "Rev growth YoY %"),
    ("rev_yoy_2", "Rev growth prior yr %"),
    ("rev_cagr_3y", "Rev CAGR %"),
    ("rev_trend", "Revenue trend"),
    ("returnOnEquity_pct", "ROE %"),
    ("profitMargins_pct", "Net margin %"),
    ("beta", "Beta"),
    ("volatility_1y_pct", "Volatility 1y %"),
    ("max_drawdown_1y_pct", "Max drawdown 1y %"),
    ("return_1y_pct", "Return 1y %"),
    ("dividendYield_pct", "Div yield %"),
    ("size_rank", "Size rank"),
    ("vanguard_weight", "VWRP weight %"),
    ("marketCap_bn", "Mkt cap ($bn)"),
    ("data_coverage_pct", "Data coverage %"),
    ("flags", "Flags"),
]


def to_dataframe(rows):
    for r in rows:
        mc = r.get("marketCap_usd") or r.get("marketCap")
        r["marketCap_bn"] = round(mc / 1e9, 1) if mc else None
        roe = r.get("returnOnEquity")
        r["returnOnEquity_pct"] = round(roe * 100, 1) if roe is not None else None
        pm = r.get("profitMargins")
        r["profitMargins_pct"] = round(pm * 100, 1) if pm is not None else None
        dy = r.get("dividendYield")
        if dy is not None:
            # yfinance returns this as a percent for most listings
            r["dividendYield_pct"] = round(dy if dy > 1 else dy * 100, 2)
        else:
            r["dividendYield_pct"] = None
        for k in ("price", "day_change_pct", "trailingPE", "forwardPE",
                  "priceToBook", "debtToEquity",
                  "net_debt_to_ebitda", "currentRatio", "beta", "rev_yoy_1",
                  "rev_yoy_2", "rev_cagr_3y", "volatility_1y_pct",
                  "max_drawdown_1y_pct", "return_1y_pct"):
            v = r.get(k)
            r[k] = round(v, 2) if isinstance(v, (int, float)) else None

    df = pd.DataFrame(rows)
    df = df[[c for c, _ in COLUMNS if c in df.columns]]
    df.columns = [label for c, label in COLUMNS if c in df.columns]
    return df


def write_excel(df, path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Screen", index=False)
        ws = writer.sheets["Screen"]
        ws.freeze_panes = "D2"
        ws.auto_filter.ref = ws.dimensions

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

        head_fill = PatternFill("solid", fgColor="1F3864")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = head_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[1].height = 34

        widths = {"Rank": 6, "Ticker": 11, "Company": 32, "Sector": 20,
                  "VWRP weight %": 13, "Opportunity score": 13,
                  "Opportunity": 12, "Size rank": 9,
                  "Region": 8, "Revenue trend": 14, "Risk rating": 12,
                  "Flags": 46}
        for i, col in enumerate(df.columns, start=1):
            letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[letter].width = widths.get(col, 13)

        n = len(df) + 1
        cols = {c: ws.cell(row=1, column=i).column_letter
                for i, c in enumerate(df.columns, start=1)}

        def rng(col):
            return f"{cols[col]}2:{cols[col]}{n}"

        if "Risk score" in cols:
            ws.conditional_formatting.add(rng("Risk score"), ColorScaleRule(
                start_type="num", start_value=20, start_color="63BE7B",
                mid_type="num", mid_value=50, mid_color="FFEB84",
                end_type="num", end_value=75, end_color="F8696B"))
        if "Risk rating" in cols:
            for text, colour in (("Low", "C6EFCE"), ("Moderate", "FFEB9C"),
                                 ("Elevated", "FFD8A8"), ("High", "FFC7CE")):
                ws.conditional_formatting.add(rng("Risk rating"), CellIsRule(
                    operator="equal", formula=[f'"{text}"'],
                    fill=PatternFill("solid", fgColor=colour)))
        if "Opportunity score" in cols:
            ws.conditional_formatting.add(rng("Opportunity score"), ColorScaleRule(
                start_type="num", start_value=35, start_color="F8696B",
                mid_type="num", mid_value=50, mid_color="FFEB84",
                end_type="num", end_value=68, end_color="63BE7B"))
        if "Opportunity" in cols:
            for text, colour in (("Strong", "C6EFCE"), ("Good", "DDF0D8"),
                                 ("Fair", "FFEB9C"), ("Weak", "FFC7CE")):
                ws.conditional_formatting.add(rng("Opportunity"), CellIsRule(
                    operator="equal", formula=[f'"{text}"'],
                    fill=PatternFill("solid", fgColor=colour)))
        if "Debt/Equity %" in cols:
            ws.conditional_formatting.add(rng("Debt/Equity %"), CellIsRule(
                operator="greaterThan", formula=["150"],
                fill=PatternFill("solid", fgColor="FFC7CE")))
        if "Rev growth YoY %" in cols:
            ws.conditional_formatting.add(rng("Rev growth YoY %"), CellIsRule(
                operator="lessThan", formula=["0"],
                fill=PatternFill("solid", fgColor="FFC7CE")))
        if "P/E (trailing)" in cols:
            ws.conditional_formatting.add(rng("P/E (trailing)"), CellIsRule(
                operator="greaterThan", formula=["40"],
                fill=PatternFill("solid", fgColor="FFEB9C")))


def write_artifact(full_html, path):
    """Same page, minus the document wrapper, for publishing as an Artifact.

    Artifacts supply their own doctype/head/body, so we hand over just the
    title, styles and content."""
    body = full_html
    for tag in ("<!doctype html>", '<html lang="en">', "<head>", "</head>",
                "<body>", "</body>", "</html>"):
        body = body.replace(tag, "")
    # Drop the meta tags the Artifact host provides itself.
    body = re.sub(r"<meta[^>]*>", "", body)
    with open(path, "w") as fh:
        fh.write(body.strip() + "\n")


DATASET_FIELDS = [
    ("ticker", "ticker"), ("name", "name"), ("region", "region"),
    ("sector", "sector"), ("currency", "currency"),
    ("vanguard_weight", "vanguardWeight"), ("marketCap_usd", "marketCapUsd"),
    ("trailingPE", "peTrailing"), ("forwardPE", "peForward"),
    ("priceToBook", "priceToBook"), ("debtToEquity", "debtToEquity"),
    ("rev_yoy_1", "revGrowth"), ("rev_cagr_3y", "revCagr"),
    ("returnOnEquity", "roe"), ("profitMargins", "netMargin"),
    ("return_1y_pct", "return1y"), ("volatility_1y_pct", "volatility1y"),
    ("max_drawdown_1y_pct", "maxDrawdown1y"),
    ("opportunity_score", "opportunity"), ("risk_score", "risk"),
]


def write_dataset(rows, path):
    """Every scored holding, for other sites to aggregate."""
    out = []
    for r in rows:
        rec = {}
        for src, dst in DATASET_FIELDS:
            v = r.get(src)
            if isinstance(v, float):
                v = round(v, 6)
            rec[dst] = v
        out.append(rec)
    payload = {
        "generated": datetime.now().strftime("%Y-%m-%d"),
        "generatedUtc": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "count": len(out),
        "source": "Yahoo Finance via vwrp-screener",
        "holdings": out,
    }
    with open(path, "w") as fh:
        json.dump(payload, fh, separators=(",", ":"))


def write_dashboard(df, path, meta):
    records = json.loads(df.to_json(orient="records"))
    payload = json.dumps({"rows": records, "meta": meta})
    html = DASHBOARD_TEMPLATE.replace("__PAYLOAD__", payload)
    with open(path, "w") as fh:
        fh.write(html)
    write_artifact(html, os.path.join(os.path.dirname(path), "artifact.html"))


DASHBOARD_TEMPLATE = r"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>VWRP Screener</title>
<link rel="manifest" href="manifest.webmanifest">
<link rel="apple-touch-icon" href="apple-touch-icon.png">
<link rel="icon" href="icon-192.png" type="image/png">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="Daily update">
<meta name="theme-color" content="#8A6A2F">
<style>
:root{
  --bg:#FAF9F6; --panel:#FFFFFF; --panel2:#F4F2EC;
  --ink:#181B22; --muted:#6B6F7B; --line:#E4E0D7; --line-strong:#CFC9BB;
  --accent:#8A6A2F; --accent-soft:#F0E8D8; --on-accent:#FFFFFF;
  --pos:#2C6E49; --neg:#A32C22;
  --s-strong:#2C6E49; --s-strong-bg:#E3EFE7;
  --s-good:#4F7C5D;  --s-good-bg:#ECF2ED;
  --s-fair:#8A6A2F;  --s-fair-bg:#F5EDDD;
  --s-weak:#A32C22;  --s-weak-bg:#F8E5E2;
  --font-display:ui-serif,Georgia,"Iowan Old Style","Times New Roman",serif;
  --font-ui:system-ui,-apple-system,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
  --font-num:ui-monospace,SFMono-Regular,"SF Mono",Menlo,Consolas,monospace;
}
@media (prefers-color-scheme:dark){:root:not([data-theme=light]){
  --bg:#0D1524; --panel:#132033; --panel2:#1A2A41;
  --ink:#E4EBF5; --muted:#8B9BB4; --line:#22334C; --line-strong:#31465F;
  --accent:#D2A85E; --accent-soft:#2A2418; --on-accent:#0D1524;
  --pos:#5FCB97; --neg:#F0857C;
  --s-strong:#5FCB97; --s-strong-bg:#10331F;
  --s-good:#8CC7A8;  --s-good-bg:#152A22;
  --s-fair:#E0BC6E;  --s-fair-bg:#2E2916;
  --s-weak:#F0857C;  --s-weak-bg:#361B1C;
}}
:root[data-theme=dark]{
  --bg:#0D1524; --panel:#132033; --panel2:#1A2A41;
  --ink:#E4EBF5; --muted:#8B9BB4; --line:#22334C; --line-strong:#31465F;
  --accent:#D2A85E; --accent-soft:#2A2418; --on-accent:#0D1524;
  --pos:#5FCB97; --neg:#F0857C;
  --s-strong:#5FCB97; --s-strong-bg:#10331F;
  --s-good:#8CC7A8;  --s-good-bg:#152A22;
  --s-fair:#E0BC6E;  --s-fair-bg:#2E2916;
  --s-weak:#F0857C;  --s-weak-bg:#361B1C;
}
*{box-sizing:border-box}
/* The hidden attribute is display:none in the UA stylesheet only, so any
   author rule setting display (.ctl is flex) silently overrides it. */
[hidden]{display:none!important}
body{margin:0;background:var(--bg);color:var(--ink);font-family:var(--font-ui);
  font-size:14px;line-height:1.45;-webkit-font-smoothing:antialiased}
.wrap{max-width:1600px;margin:0 auto;padding:34px 22px 64px;position:relative}
.themebtn{position:absolute;top:32px;right:22px;background:var(--panel);
  color:var(--muted);border:1px solid var(--line-strong);border-radius:2px;
  padding:6px 12px;font-size:11px;letter-spacing:.06em;text-transform:uppercase;
  cursor:pointer;font-weight:600;font-family:var(--font-ui)}
.themebtn:hover{color:var(--accent);border-color:var(--accent)}
.eyebrow{font-size:10.5px;letter-spacing:.18em;text-transform:uppercase;
  color:var(--accent);font-weight:600;margin-bottom:10px}
h1{font-family:var(--font-display);font-size:clamp(28px,4vw,40px);font-weight:600;
  margin:0;letter-spacing:-.015em;text-wrap:balance}
.sub{color:var(--muted);font-size:13px;margin-top:8px;padding-bottom:22px;
  border-bottom:1px solid var(--line-strong)}
.livebar{display:flex;align-items:center;gap:8px;margin-top:14px;font-size:11.5px;
  color:var(--muted);letter-spacing:.02em}
.dot{width:7px;height:7px;border-radius:50%;background:var(--muted);flex:none}
.dot.live{background:var(--pos);box-shadow:0 0 0 3px color-mix(in srgb,var(--pos) 22%,transparent)}
.dot.stale{background:var(--s-fair)}
.cards{display:grid;grid-template-columns:repeat(4,1fr);
  gap:1px;margin:22px 0;background:var(--line);border:1px solid var(--line);
  border-radius:3px;overflow:hidden}
.card{background:var(--panel);padding:15px 16px}
.card .k{font-size:10px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted)}
.card .v{font-family:var(--font-display);font-size:25px;font-weight:600;
  margin-top:5px;letter-spacing:-.02em;font-variant-numeric:tabular-nums}
.controls{background:var(--panel);border:1px solid var(--line);border-radius:3px;
  padding:15px 16px;margin-bottom:16px;display:flex;flex-wrap:wrap;gap:14px;align-items:flex-end}
.ctl{display:flex;flex-direction:column;gap:5px}
.ctl label{font-size:10px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted)}
input,select{background:var(--bg);color:var(--ink);border:1px solid var(--line-strong);
  border-radius:2px;padding:7px 9px;font-size:13px;min-width:132px;font-family:inherit}
input[type=search]{min-width:200px}
input:focus-visible,select:focus-visible,button:focus-visible{outline:2px solid var(--accent);
  outline-offset:1px}
button{background:var(--accent);color:var(--on-accent);border:0;border-radius:2px;padding:8px 15px;
  font-size:12px;letter-spacing:.04em;cursor:pointer;font-family:inherit;font-weight:600}
button.ghost{background:transparent;color:var(--muted);border:1px solid var(--line-strong);
  font-weight:500}
button:disabled{opacity:.55;cursor:progress}
.count{color:var(--muted);font-size:11.5px;margin:0 0 8px;letter-spacing:.02em}
.tablewrap{background:var(--panel);border:1px solid var(--line);border-radius:3px;
  overflow:auto;max-height:74vh}
table{border-collapse:separate;border-spacing:0;width:100%;font-size:12.5px;white-space:nowrap}
th{position:sticky;top:0;background:var(--panel);border-bottom:1.5px solid var(--line-strong);
  padding:11px 10px;text-align:right;font-size:10px;text-transform:uppercase;
  letter-spacing:.08em;color:var(--muted);cursor:pointer;user-select:none;z-index:2;font-weight:600}
th:first-child,th:nth-child(2),th:nth-child(3){text-align:left}
th:hover{color:var(--accent)}
th.sorted{color:var(--accent)}
th.sorted::after{content:" \25BC";font-size:8px}
th.sorted.asc::after{content:" \25B2"}
td{padding:8px 10px;border-bottom:1px solid var(--line);text-align:right;
  font-family:var(--font-num);font-size:12px;font-variant-numeric:tabular-nums}
td:first-child,td:nth-child(2),td:nth-child(3){text-align:left}
td.name{font-family:var(--font-ui);white-space:normal;min-width:180px;max-width:250px}
td.txt{font-family:var(--font-ui);text-align:right}
td.flags{font-family:var(--font-ui);white-space:normal;min-width:170px;max-width:270px;
  color:var(--muted);font-size:11.5px;text-align:left}
tbody tr:hover td{background:var(--panel2)}
.pill{display:inline-block;padding:2px 9px;border-radius:2px;font-size:10.5px;
  font-weight:600;letter-spacing:.05em;text-transform:uppercase;font-family:var(--font-ui)}
.Low,.Strong{background:var(--s-strong-bg);color:var(--s-strong)}
.Good{background:var(--s-good-bg);color:var(--s-good)}
.Moderate,.Fair{background:var(--s-fair-bg);color:var(--s-fair)}
.Elevated{background:var(--s-fair-bg);color:var(--s-fair);opacity:.85}
.High,.Weak{background:var(--s-weak-bg);color:var(--s-weak)}
.pos{color:var(--pos)}.neg{color:var(--neg)}
.gloss{margin-top:22px;background:var(--panel);border:1px solid var(--line);
  border-radius:3px;padding:15px 17px}
.gloss summary{cursor:pointer;font-weight:600;font-size:13px;font-family:var(--font-display)}
.gloss summary:hover{color:var(--accent)}
.gloss dl{margin:16px 0 0;display:grid;grid-template-columns:minmax(150px,215px) 1fr;
  gap:11px 20px}
.gloss dt{font-weight:600;font-size:12.5px}
.gloss dd{margin:0;color:var(--muted);font-size:12.5px;line-height:1.55;max-width:68ch}
@media(max-width:820px){.cards{grid-template-columns:repeat(2,1fr)}}
@media(max-width:700px){.gloss dl{grid-template-columns:1fr;gap:3px 0}
  .gloss dd{margin-bottom:11px}}
.note{margin-top:16px;color:var(--muted);font-size:11.5px;line-height:1.65;
  background:var(--panel);border:1px solid var(--line);border-left:2px solid var(--accent);
  border-radius:3px;padding:14px 17px;max-width:92ch}
@media(prefers-reduced-motion:reduce){*{transition:none!important;animation:none!important}}
</style></head><body><div class="wrap">
<button id="theme" class="themebtn" type="button" aria-label="Switch theme"></button>
<div class="eyebrow">FTSE All-World &middot; VWRP</div>
<h1>VWRP Stock Screener</h1>
<div class="sub" id="sub"></div>
<div class="livebar"><span class="dot" id="dot"></span><span id="livetext"></span></div>
<div class="cards" id="cards"></div>
<div class="controls">
  <div class="ctl"><label>Search</label><input type="search" id="q" placeholder="Company or ticker"></div>
  <div class="ctl"><label>Sector</label><select id="sector"><option value="">All</option></select></div>
  <div class="ctl"><label>Region</label><select id="region"><option value="">All</option></select></div>
  <div class="ctl"><label>Risk rating</label><select id="risk"><option value="">All</option>
    <option>Low</option><option>Moderate</option><option>Elevated</option><option>High</option></select></div>
  <div class="ctl"><label>Max P/E</label><input type="number" id="maxpe" placeholder="any"></div>
  <div class="ctl"><label>Max debt/equity %</label><input type="number" id="maxde" placeholder="any"></div>
  <div class="ctl"><label>Min rev growth %</label><input type="number" id="minrev" placeholder="any"></div>
  <div class="ctl"><label>&nbsp;</label><button class="ghost" id="reset">Reset</button></div>
  <div class="ctl" id="refreshwrap" hidden><label>&nbsp;</label>
    <button id="refresh">Refresh data</button></div>
</div>
<div class="count" id="count"></div>
<div class="tablewrap"><table><thead><tr id="head"></tr></thead><tbody id="body"></tbody></table></div>
<details class="gloss"><summary>What every column means</summary>
<dl id="glossary"></dl></details>
<div class="note" id="note"></div>
</div>
<script>
const DATA = __PAYLOAD__;
const rows = DATA.rows, meta = DATA.meta;
const COLS = Object.keys(rows[0] || {});
const NUMERIC = new Set(COLS.filter(c => rows.some(r => typeof r[c] === "number")));
const SIGNED = new Set(["Rev growth YoY %","Rev growth prior yr %","Rev CAGR %","Return 1y %","Max drawdown 1y %","Day %"]);
let sortCol = "Rank", sortAsc = true;

function fmt(v, col){
  if(v === null || v === undefined || v === "") return "–";
  if(typeof v !== "number") return v;
  if(col === "Mkt cap ($bn)") return v.toLocaleString(undefined,{maximumFractionDigits:0});
  if(Number.isInteger(v)) return v;
  return v.toFixed(Math.abs(v) < 10 ? 2 : 1);
}
function uniq(col){ return [...new Set(rows.map(r=>r[col]).filter(Boolean))].sort(); }
for(const v of uniq("Sector")) sector.add(new Option(v,v));
for(const v of uniq("Region")) region.add(new Option(v,v));

document.getElementById("head").innerHTML = COLS.map(c=>`<th data-c="${c}">${c}</th>`).join("");
document.querySelectorAll("th").forEach(th=>th.onclick=()=>{
  const c = th.dataset.c;
  if(sortCol===c) sortAsc=!sortAsc; else {sortCol=c; sortAsc=!NUMERIC.has(c)||c==="Rank"||c==="Risk score";}
  render();
});

function filtered(){
  const q = document.getElementById("q").value.toLowerCase().trim();
  const s = sector.value, rg = region.value, rk = risk.value;
  const mpe = parseFloat(maxpe.value), mde = parseFloat(maxde.value), mrv = parseFloat(minrev.value);
  return rows.filter(r=>{
    if(q && !((r["Company"]||"")+" "+(r["Ticker"]||"")).toLowerCase().includes(q)) return false;
    if(s && r["Sector"]!==s) return false;
    if(rg && r["Region"]!==rg) return false;
    if(rk && r["Risk rating"]!==rk) return false;
    if(!isNaN(mpe)){ const p=r["P/E (trailing)"]; if(p===null||p===undefined||p>mpe||p<0) return false; }
    if(!isNaN(mde)){ const d=r["Debt/Equity %"]; if(d===null||d===undefined||d>mde) return false; }
    if(!isNaN(mrv)){ const g=r["Rev growth YoY %"]; if(g===null||g===undefined||g<mrv) return false; }
    return true;
  });
}

function render(){
  const data = filtered().slice().sort((a,b)=>{
    let x=a[sortCol], y=b[sortCol];
    if(x===null||x===undefined) return 1;
    if(y===null||y===undefined) return -1;
    if(typeof x==="number") return sortAsc? x-y : y-x;
    return sortAsc? String(x).localeCompare(String(y)) : String(y).localeCompare(String(x));
  });
  document.getElementById("body").innerHTML = data.map(r=>"<tr>"+COLS.map(c=>{
    const v = r[c];
    if((c==="Risk rating"||c==="Opportunity") && v) return `<td><span class="pill ${v}">${v}</span></td>`;
    if(c==="Company") return `<td class="name">${v??"–"}</td>`;
    if(c==="Sector"||c==="Region"||c==="Revenue trend") return `<td class="txt">${v??"–"}</td>`;
    if(c==="Flags") return `<td class="flags">${v||""}</td>`;
    let cls="";
    if(SIGNED.has(c) && typeof v==="number") cls = v<0 ? "neg" : (v>0?"pos":"");
    return `<td class="${cls}">${fmt(v,c)}</td>`;
  }).join("")+"</tr>").join("");
  document.getElementById("count").textContent =
    `Showing ${data.length} of ${rows.length} holdings`;
  document.querySelectorAll("th").forEach(th=>{
    th.classList.toggle("sorted", th.dataset.c===sortCol);
    th.classList.toggle("asc", th.dataset.c===sortCol && sortAsc);
  });
  stats(data);
}

function median(a){ if(!a.length) return null; const s=a.slice().sort((x,y)=>x-y);
  const m=Math.floor(s.length/2); return s.length%2? s[m] : (s[m-1]+s[m])/2; }
function col(data,c){ return data.map(r=>r[c]).filter(v=>typeof v==="number"); }

function stats(data){
  const pe = median(col(data,"P/E (trailing)").filter(v=>v>0));
  const de = median(col(data,"Debt/Equity %"));
  const rv = median(col(data,"Rev growth YoY %"));
  const counts = {Low:0,Moderate:0,Elevated:0,High:0};
  data.forEach(r=>{ if(counts[r["Risk rating"]]!==undefined) counts[r["Risk rating"]]++; });
  const cards = [
    ["Holdings shown", data.length],
    ["Median P/E", pe===null?"–":pe.toFixed(1)],
    ["Median debt/equity", de===null?"–":de.toFixed(0)+"%"],
    ["Median rev growth", rv===null?"–":rv.toFixed(1)+"%"],
    ["Low risk", counts.Low],
    ["Moderate", counts.Moderate],
    ["Elevated", counts.Elevated],
    ["High risk", counts.High],
  ];
  document.getElementById("cards").innerHTML = cards.map(([k,v])=>
    `<div class="card"><div class="k">${k}</div><div class="v">${v}</div></div>`).join("");
}

["q","sector","region","risk","maxpe","maxde","minrev"].forEach(id=>{
  const el=document.getElementById(id);
  el.addEventListener(el.tagName==="SELECT"?"change":"input", render);
});
document.getElementById("reset").onclick=()=>{
  ["q","maxpe","maxde","minrev"].forEach(id=>document.getElementById(id).value="");
  ["sector","region","risk"].forEach(id=>document.getElementById(id).value="");
  render();
};
// Theme: remembered per browser, "auto" follows the operating system.
// localStorage throws outright in sandboxed contexts (private browsing, some
// embedded viewers), so every access is guarded -- an unguarded one here would
// take the whole script down and leave the page with no table at all.
const store = {
  get(k){ try { return localStorage.getItem(k); } catch(e){ return null; } },
  set(k,v){ try { localStorage.setItem(k,v); } catch(e){ /* not persisted */ } }
};
(function(){
  const btn=document.getElementById("theme"), root=document.documentElement;
  if(!btn) return;
  const LABEL={auto:"Auto",light:"Light",dark:"Dark"};
  let mode=store.get("vwrp-theme")||"auto";
  if(!LABEL[mode]) mode="auto";
  function apply(){
    if(mode==="auto") root.removeAttribute("data-theme");
    else root.setAttribute("data-theme",mode);
    btn.textContent=LABEL[mode];
    store.set("vwrp-theme",mode);
  }
  btn.onclick=()=>{ mode = mode==="auto"?"dark":mode==="dark"?"light":"auto"; apply(); };
  apply();
})();

const GLOSSARY = [
 ["Rank","Position in this screen, best opportunity first. It is the ranking on Opportunity score, not company size \u2014 so a small holding with strong numbers outranks a mega-cap with weak ones."],
 ["Ticker","Yahoo Finance symbol. The suffix marks the exchange: no suffix is the US, .L London, .T Tokyo, .PA Paris, .DE Germany, .HK Hong Kong, .TW Taiwan, .KS Korea, .NS India, .AX Australia, .TO Toronto, .SW Switzerland."],
 ["Company","The company\u2019s registered name as reported by the exchange."],
 ["Opportunity score","0\u2013100, higher is better. A blend of value 30%, growth 30%, quality 25% and stability 15%. It rewards cheap multiples, so it surfaces value traps and cyclicals at peak earnings next to genuine bargains \u2014 always read Flags and Risk rating alongside it."],
 ["Opportunity","The score bucketed: Strong 62+, Good 52\u201361, Fair 42\u201351, Weak below 42."],
 ["Risk score","0\u2013100, LOWER is safer \u2014 the opposite direction to Opportunity. Blends leverage 25%, volatility 25%, valuation 20%, growth quality 20%, profitability 10%."],
 ["Risk rating","The risk score bucketed: Low under 35, Moderate 35\u201349, Elevated 50\u201364, High 65+."],
 ["Sector","Broad industry grouping. Useful because multiples are only comparable within a sector \u2014 a utility on 18x is expensive, a software firm on 18x is cheap."],
 ["Region","Where the shares are listed, not necessarily where the company earns its money."],
 ["P/E (trailing)","Share price divided by the last 12 months of earnings per share. How many years of current profit you pay for one share. Blank means the company lost money, so the ratio is meaningless."],
 ["P/E (forward)","The same ratio using analysts\u2019 forecast earnings for the year ahead. Lower than trailing means profits are expected to grow; higher means they are expected to fall."],
 ["P/B","Price to book \u2014 market value against balance-sheet net assets. Below 1 means the market values the firm at less than its accounting net worth. Most meaningful for banks and insurers, least for asset-light software."],
 ["Debt/Equity %","Total borrowings as a percentage of shareholder equity. Under 50% is conservative, over 100% means more debt than equity, over 200% is aggressive. Ignore it for banks \u2014 borrowing is their raw material, so the figure is often blank or meaningless."],
 ["Net debt/EBITDA","Borrowings minus cash, divided by annual cash earnings \u2014 roughly how many years of earnings it would take to repay debt. Under 2x is comfortable, over 3x is stretched, over 4x limits what a company can do in a downturn."],
 ["Current ratio","Short-term assets divided by short-term liabilities. Above 1 means the company can cover the coming year\u2019s bills from liquid assets; below 1 means it depends on refinancing or incoming cash flow."],
 ["Rev growth YoY %","Revenue in the most recent reported year against the year before. Sales growth, before any cost or accounting effects."],
 ["Rev growth prior yr %","The same figure one year earlier. Read it against the column to its left: that comparison is what tells you whether growth is speeding up or slowing down."],
 ["Rev CAGR %","Compound annual revenue growth across all the years available, usually four. Smooths out a single freak year that a one-year figure would exaggerate."],
 ["Revenue trend","Plain-English read of the last two growth rates. Accelerating: growth faster than last year. Steady: broadly unchanged. Decelerating: still growing but slower. Recovering: back to growth after a decline. Contracting: sales falling."],
 ["ROE %","Return on equity \u2014 annual profit as a percentage of shareholder capital. How hard the company works the money invested in it. Above 15% is good, above 25% is exceptional, but very high figures often mean heavy debt rather than excellence."],
 ["Net margin %","Profit kept as a percentage of every pound of sales. Thin margins mean little cushion when costs rise; fat margins usually signal pricing power, and attract competitors."],
 ["Beta","Volatility relative to the market. 1.0 moves with the market, 1.5 amplifies moves by half, below 1.0 is steadier than average, negative means it has tended to move opposite the market."],
 ["Volatility 1y %","How much the share price has actually swung over the past year, annualised. Unlike beta this measures real movement rather than correlation. Under 20% is calm, over 40% is turbulent."],
 ["Max drawdown 1y %","The worst peak-to-trough fall over the past year \u2014 what you would have lost buying at the top and selling at the bottom. The most honest single measure of how uncomfortable holding it has been."],
 ["Return 1y %","Total share price change over the past year, dividends excluded. Context only: it says where the price has been, not where it is going."],
 ["Div yield %","Annual dividend as a percentage of the share price. An unusually high yield is often the arithmetic of a falling share price rather than generosity, and can precede a cut."],
 ["Size rank","Position by market value among the holdings shown, largest first. This was the old default ordering, kept as context."],
 ["Weight % of set","Share of the holdings shown here, not of the whole fund. VWRP holds ~3,750 stocks, so roughly halve this for the true weight inside the ETF."],
 ["Mkt cap ($bn)","Total value of all shares, converted to US dollars at live spot exchange rates."],
 ["Data coverage %","How much of the risk model had data for this stock. 100% means every input was present; 75% is common for banks, where debt ratios do not apply. Treat scores below 75% as rough."],
 ["Flags","Automatic warnings from the numbers \u2014 high debt, contracting revenue, negative free cash flow, loss-making, rich valuation. An empty cell means nothing tripped."],
];
document.getElementById("glossary").innerHTML = GLOSSARY.map(([k,v])=>
  `<dt>${k}</dt><dd>${v}</dd>`).join("");
// Live quotes: only available when served by serve.py, which proxies Yahoo.
// Opened as a plain file, the page stays an honest snapshot.
const byTicker = new Map(rows.map(r=>[r["Ticker"], r]));
let liveTimer = null;

function setLive(state, text){
  const dot=document.getElementById("dot");
  dot.className = "dot" + (state ? " " + state : "");
  document.getElementById("livetext").textContent = text;
}

async function pollQuotes(){
  try{
    const res = await fetch("/quotes");
    if(!res.ok) throw new Error("quote fetch failed");
    const data = await res.json();
    let n = 0;
    for(const [tk, q] of Object.entries(data.quotes || {})){
      const row = byTicker.get(tk);
      if(!row || q.price == null) continue;
      row["Price"] = q.price;
      if(q.day_pct != null) row["Day %"] = q.day_pct;
      n++;
    }
    render();
    if(n === 0){
      setLive("stale", "Quote feed returned nothing just now — retrying in 60s. "+
        "Prices shown are from the last run.");
      return;
    }
    const t = new Date(data.asof).toLocaleTimeString();
    setLive("live", `Live prices · ${n} quotes updated ${t} · refreshes every 60s · `+
      `exchange feeds are typically delayed up to 15 minutes`);
  }catch(err){
    // No /quotes endpoint. On localhost that means the feed broke; anywhere
    // else this is the published copy, which is a daily snapshot by design.
    // Only the GitHub Pages copy actually rebuilds on a schedule. Everywhere
    // else -- a shared claude.ai link, a saved file -- it is a fixed snapshot,
    // and must not claim otherwise.
    const host = location.hostname;
    const isLocal = isLocalHost(host);
    const rebuilds = host.endsWith(".github.io");
    setLive("", isLocal
      ? `Live prices unavailable right now — showing prices as of ${meta.generated}.`
      : rebuilds
        ? `Snapshot — prices as of ${meta.generated}. This page rebuilds `+
          `automatically each evening after the US close.`
        : `Fixed snapshot — prices as of ${meta.generated}. This copy does not `+
          `update on its own.`);
    clearInterval(liveTimer); liveTimer = null;
  }
}

if(location.protocol.startsWith("http")){
  setLive("", "Connecting to live prices\u2026");
  pollQuotes();
  liveTimer = setInterval(pollQuotes, 60000);
  // Don't poll a tab nobody is looking at.
  document.addEventListener("visibilitychange", ()=>{
    if(document.hidden){ clearInterval(liveTimer); liveTimer=null; }
    else if(!liveTimer){ pollQuotes(); liveTimer=setInterval(pollQuotes,60000); }
  });
}else{
  setLive("", `Snapshot — prices as of ${meta.generated}. `+
    `Run \u201cpython serve.py\u201d for live updating prices.`);
}

// The refresh endpoint only exists when the page is served by serve.py.
// Refresh re-runs the screener through serve.py, so it exists only where that
// server is reachable: this machine, or another device on the same network.
// On the published site the button would just fail for every visitor.
function isLocalHost(h){
  return h === "localhost" || h === "127.0.0.1" ||
         /^192\.168\./.test(h) || /^10\./.test(h) ||
         /^172\.(1[6-9]|2\d|3[01])\./.test(h);
}
if(isLocalHost(location.hostname)){
  const wrap=document.getElementById("refreshwrap"); wrap.hidden=false;
  const btn=document.getElementById("refresh");
  btn.onclick=async()=>{
    btn.disabled=true; const original=btn.textContent;
    btn.textContent="Refreshing\u2026 this can take a few minutes";
    try{
      const res=await fetch("/refresh",{method:"POST"});
      if(!res.ok) throw new Error(await res.text());
      location.reload();
    }catch(err){
      btn.textContent="Refresh failed \u2014 check the terminal";
      setTimeout(()=>{btn.textContent=original; btn.disabled=false;},4000);
    }
  };
}
document.getElementById("sub").textContent =
  `${meta.count} largest holdings, ranked best opportunity first · data as of `+
  `${meta.generated} · source: Yahoo Finance`;
document.getElementById("note").innerHTML =
  `<b>Risk score</b> (0–100, lower is safer) blends leverage 25%, volatility 25%, `+
  `valuation 20%, revenue-growth quality 20%, profitability 10%. Weights are renormalised `+
  `over whichever inputs have data — check the <i>Data coverage %</i> column before `+
  `leaning on a score. <b>Approx VWRP wt %</b> is derived from live market caps, not `+
  `Vanguard's published free-float weights, so treat it as an approximation. `+
  `Market caps are converted to USD at live spot FX. P/E, debt/equity, growth and `+
  `margin figures are ratios, so they are currency-neutral by construction. `+
  `Not investment advice.`;
render();
</script></body></html>
"""


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Screen the largest VWRP holdings.")
    ap.add_argument("--top", type=int, default=150, help="how many holdings (default 150)")
    ap.add_argument("--refresh", action="store_true", help="ignore cache")
    ap.add_argument("--max-pe", type=float, help="only keep stocks with trailing P/E below this")
    ap.add_argument("--max-de", type=float, help="only keep stocks with debt/equity below this")
    ap.add_argument("--min-growth", type=float, help="only keep stocks with rev growth above this %%")
    ap.add_argument("--pool", type=int, default=0,
                    help="cap on holdings fetched; 0 = the whole file")
    ap.add_argument("--workers", type=int, default=5)
    ap.add_argument("--fetch-only", help=argparse.SUPPRESS)
    args = ap.parse_args()

    # Internal mode: warm the cache for specific tickers in a clean process.
    if args.fetch_only:
        for tk in [t for t in args.fetch_only.split(",") if t]:
            rec = fetch_one(tk, use_cache=False)
            print(f"  {tk}: {'ok' if rec.get('marketCap') else 'still missing'}",
                  file=sys.stderr)
            time.sleep(0.8)
        return None

    os.makedirs(OUTDIR, exist_ok=True)
    if os.path.exists(HOLDINGS):
        universe = load_holdings(args.pool)
        total = sum(1 for l in open(HOLDINGS)
                    if l.strip() and not l.startswith(("#", "ticker,")))
        covered = sum(u["vanguard_weight"] for u in universe)
        print(f"VWRP holdings on file: {total} | fetching heaviest "
              f"{len(universe)} ({covered:.1f}% of fund)", file=sys.stderr)
    else:
        universe = load_universe()
        print(f"Universe: {len(universe)} candidate tickers", file=sys.stderr)

    print("Fetching fundamentals (cached for 24h)...", file=sys.stderr)
    records = fetch_all(universe, use_cache=not args.refresh, workers=args.workers)

    # Yahoo throttles bursts by silently returning partial quote payloads, and
    # once a process is throttled it stays throttled -- in-process retries never
    # recover. Sweep the misses in fresh subprocesses, which repopulate the
    # cache, then re-read it here.
    for rnd in (1, 2, 3):
        misses = [r for r in records if not r.get("marketCap")
                  and "Quote not found" not in (r.get("error") or "")]
        if not misses:
            break
        print(f"Retry round {rnd}: {len(misses)} tickers in a fresh process...",
              file=sys.stderr)
        time.sleep(4)
        subprocess.run(
            [sys.executable, "-W", "ignore", os.path.abspath(__file__),
             "--fetch-only", ",".join(r["ticker"] for r in misses)],
            check=False)
        fixed = {}
        for r in misses:
            cached = read_cache(r["ticker"])
            if cached and cached.get("marketCap"):
                cached["region"] = r["region"]
                fixed[r["ticker"]] = cached
        print(f"  recovered {len(fixed)}/{len(misses)}", file=sys.stderr)
        if not fixed:
            break
        records = [fixed.get(r["ticker"], r) for r in records]

    ok = [r for r in records if r.get("marketCap")]
    failed = [r for r in records if not r.get("marketCap")]

    print("Fetching FX rates to normalise market caps to USD...", file=sys.stderr)
    fx = fetch_fx({r.get("currency") for r in ok})
    for r in ok:
        rate = fx.get(r.get("currency") or "USD", 1.0)
        r["marketCap_usd"] = r["marketCap"] * rate
        r["fx_rate"] = rate

    ok.sort(key=lambda r: r["marketCap_usd"], reverse=True)

    # Dual listings (e.g. HSBC on London and Hong Kong) are one company and
    # would otherwise occupy two ranks. Keep the largest listing of each.
    seen, deduped, dropped = {}, [], []   # name -> kept record
    for r in ok:
        key = (r.get("longName") or r.get("shortName") or r["ticker"]).strip().lower()
        if key in seen:
            # Same company, second listing or share class. Keep one row but
            # carry its fund weight over, or VWRP's true exposure is understated.
            keeper = seen[key]
            if r.get("vanguard_weight") and keeper.get("vanguard_weight") is not None:
                keeper["vanguard_weight"] += r["vanguard_weight"]
            dropped.append(f"{r['ticker']} (merged into {keeper['ticker']})")
            continue
        seen[key] = r
        deduped.append(r)
    if dropped:
        print(f"Dropped {len(dropped)} dual listings: " + ", ".join(dropped[:8]) +
              (" ..." if len(dropped) > 8 else ""), file=sys.stderr)
    ok = deduped
    top = ok[:args.top]
    print(f"Resolved {len(ok)} tickers ({len(failed)} failed); keeping top {len(top)}",
          file=sys.stderr)
    if failed:
        print("  failed: " + ", ".join(r["ticker"] for r in failed[:15]) +
              (" ..." if len(failed) > 15 else ""), file=sys.stderr)

    # Score every resolved holding, not just the ones we display: the geography
    # page aggregates by country and needs the whole set to do that honestly.
    price_risk = fetch_price_risk([r["ticker"] for r in ok])
    rows = build_rows(ok, price_risk)

    total_mc = sum(r["marketCap_usd"] for r in rows)
    for i, r in enumerate(sorted(rows, key=lambda x: -x["marketCap_usd"]), start=1):
        r["size_rank"] = i
        r["approx_weight_pct"] = round(r["marketCap_usd"] / total_mc * 100, 2)

    # Best screening opportunities first; anything unscoreable sinks to the end.
    for i, r in enumerate(sorted(rows, key=lambda x: -(x.get("opportunity_score") or -1)),
                          start=1):
        r["rank"] = i

    # Optional command-line filters
    before = len(rows)
    if args.max_pe is not None:
        rows = [r for r in rows if r.get("trailingPE") is not None
                and 0 < r["trailingPE"] <= args.max_pe]
    if args.max_de is not None:
        rows = [r for r in rows if r.get("debtToEquity") is not None
                and r["debtToEquity"] <= args.max_de]
    if args.min_growth is not None:
        rows = [r for r in rows if r.get("rev_yoy_1") is not None
                and r["rev_yoy_1"] >= args.min_growth]
    if len(rows) != before:
        print(f"Filters kept {len(rows)} of {before}", file=sys.stderr)

    rows.sort(key=lambda r: r["rank"])

    # Machine-readable output for every holding, before trimming to the display
    # set. The geography site consumes this nightly.
    write_dataset(rows, os.path.join(OUTDIR, "data.json"))

    df = to_dataframe(rows[:args.top])

    stamp = datetime.now().strftime("%Y-%m-%d")
    xlsx = os.path.join(OUTDIR, f"vwrp_screen_{stamp}.xlsx")
    html = os.path.join(OUTDIR, "dashboard.html")
    write_excel(df, xlsx)
    write_dashboard(df, html, {"count": len(df), "generated": stamp})

    print(f"\nWrote {xlsx}")
    print(f"Wrote {html}")
    print(f"Wrote {os.path.join(OUTDIR, 'artifact.html')}")
    print(f"Wrote {os.path.join(OUTDIR, 'data.json')} ({len(rows)} holdings)")
    return df


if __name__ == "__main__":
    main()
